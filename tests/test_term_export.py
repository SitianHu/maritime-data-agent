from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from app import db
from app.main import build_terms_export, export_terms


class TermExportTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.db_patch = patch.object(db, "DB_PATH", Path(self.temp_dir.name) / "terms.db")
        self.db_patch.start()
        db.init_db()
        db.save_dataset("dataset-1", "实时船舶", "live_table", "live.xlsx", 0, [])
        db.add_term("活跃船舶", "当前有动态数据的船舶", "在线船舶", "dataset-1")
        db.add_term("重点目标", "需要重点关注的目标", "", None)

    def tearDown(self):
        self.db_patch.stop()
        self.temp_dir.cleanup()

    def test_export_contains_required_columns_and_dataset_names(self):
        workbook = load_workbook(build_terms_export())
        sheet = workbook["术语库"]
        self.assertEqual([cell.value for cell in sheet[1]], ["术语", "定义", "同义词", "关联数据表"])
        rows = {row[0]: row for row in sheet.iter_rows(min_row=2, values_only=True)}
        self.assertEqual(rows["活跃船舶"], ("活跃船舶", "当前有动态数据的船舶", "在线船舶", "实时船舶"))
        self.assertEqual(rows["重点目标"][3], "全局术语")
        self.assertEqual(sheet.freeze_panes, "A2")

    def test_export_response_is_an_xlsx_download(self):
        response = export_terms()
        self.assertEqual(response.media_type, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertIn("attachment;", response.headers["content-disposition"])
        self.assertTrue(response.headers["content-disposition"].endswith('.xlsx"'))


if __name__ == "__main__":
    unittest.main()
